import re
from chat_utils.openai_provider import transcribe_audio_bytes
import io
from chat_utils.db import add_user_sync
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from docx import Document
from pypdf import PdfReader
import csv
import os


def is_valid_email(email: str) -> bool:
    """
    Check if the provided email has a valid format.
    Args:
        email (str): The email address to validate.
    Returns:
        bool: True if the email format is valid, False otherwise.
    """
    EMAIL_REGEX = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
    return EMAIL_REGEX.match(email) is not None


def init_db():
    """
    Inizialize the database with admin user from env variables.
    Raises:
        ValueError: if ADMIN_USER_ID is not set in env variables."""
    if "ADMIN_USER_ID" not in os.environ:
        raise ValueError("ADMIN_USER_ID is not detected in env variables.")
    add_user_sync(
        user_id=os.environ["ADMIN_USER_ID"],
        mail=os.getenv("ADMIN_MAIL", ""),
        name=os.getenv("ADMIN_NAME", ""),
        surname=os.getenv("SURNAME_MAIL", ""),
        role="admin",
    )


async def build_content_blocks(message, context):
    """
    Build content blocks from a Telegram message.
    Supports text, images, audio (voice/audio/video-note), and documents (PDF, Excel, Word, CSV).
    Returns a list of content blocks with type and content.
    Args:
        message: Telegram message object.
        context: Telegram context object.
    Returns:
        List[Dict]: List of content blocks.
    """
    content_blocks = []

    # TEXT (normale o caption)
    user_text = message.text or message.caption
    if user_text:
        content_blocks.append({"type": "text", "text": user_text})

    # IMAGE (photo o document immagine)
    if message.photo:
        photo = message.photo[-1]
        file = await context.bot.get_file(photo.file_id)
        image_url = file.file_path
        content_blocks.append({"type": "image", "url": image_url})

    if (
        message.document
        and message.document.mime_type
        and message.document.mime_type.startswith("image/")
    ):
        file = await context.bot.get_file(message.document.file_id)
        image_url = file.file_path
        content_blocks.append({"type": "image", "url": image_url})

    # AUDIO (voice/audio/video-note)
    voice = message.voice or message.audio or message.video_note
    if voice:
        file = await context.bot.get_file(voice.file_id)
        bio = io.BytesIO()
        await file.download_to_memory(bio)
        transcript_text = transcribe_audio_bytes(bio.getvalue(), "audio.ogg")
        content_blocks.append({"type": "text", "text": transcript_text})

    # DOCUMENTS (PDF, Excel, Word, CSV)
    if message.document and not (message.document.mime_type or "").startswith("image/"):
        doc = message.document
        file = await context.bot.get_file(doc.file_id)
        bio = io.BytesIO()
        await file.download_to_memory(bio)
        file_bytes = bio.getvalue()

        extracted_text = extract_text_from_document(
            file_bytes=file_bytes,
            mime_type=doc.mime_type or "",
            file_name=doc.file_name or "",
        )

        if extracted_text:
            content_blocks.append({"type": "text", "text": extracted_text})

    return content_blocks


def extract_text_from_document(
    file_bytes: bytes, mime_type: str, file_name: str
) -> str:
    """
    Extract text from various document types based on MIME type or file extension.
    Supported types: PDF, Excel, Word, CSV.
    Args:
        file_bytes (bytes): The content of the file in bytes.
        mime_type (str): The MIME type of the file.
        file_name (str): The name of the file.
    Returns:
        str: Extracted text content."""
    mime_type = mime_type.lower()
    file_name_lower = file_name.lower()

    # PDF
    if "pdf" in mime_type or file_name_lower.endswith(".pdf"):
        return extract_text_from_pdf(file_bytes)

    # EXCEL
    if any(
        t in mime_type
        for t in [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ]
    ) or file_name_lower.endswith((".xlsx", ".xls")):
        return extract_text_from_excel(file_bytes)

    # WORD
    if any(
        t in mime_type
        for t in [
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/msword",
        ]
    ) or file_name_lower.endswith((".docx", ".doc")):
        return extract_text_from_word(file_bytes)

    # CSV
    if "csv" in mime_type or file_name_lower.endswith(".csv"):
        return extract_text_from_csv(file_bytes)

    # Fallback
    return f"Documento ricevuto: {file_name} (MIME: {mime_type}). Tipo non supportato per l'estrazione testi."


def extract_text_from_pdf(file_bytes: bytes, max_pages: int = 10) -> str:
    """
    Extract text from a PDF file.
    Args:
        file_bytes (bytes): The content of the PDF file in bytes.
        max_pages (int): Maximum number of pages to extract.
    Returns:
        str: Extracted text content."""
    bio = io.BytesIO(file_bytes)
    reader = PdfReader(bio)
    texts = []

    for i, page in enumerate(reader.pages):
        if i >= max_pages:
            break
        page_text = page.extract_text() or ""
        texts.append(page_text)

    text = "\n\n".join(texts).strip()
    if not text:
        return "PDF ricevuto ma non è stato possibile estrarre testo."
    return text


def extract_text_from_word(file_bytes: bytes) -> str:
    """
    Extract text from a Word document (.docx).
    Args:
        file_bytes (bytes): The content of the Word file in bytes.
    Returns:
        str: Extracted text content."""
    bio = io.BytesIO(file_bytes)
    doc = Document(bio)
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    if not paragraphs:
        return "Documento Word ricevuto ma vuoto o non leggibile."
    return "\n".join(paragraphs)


def extract_text_from_excel(file_bytes: bytes, max_rows: int = 100) -> str:
    """
    Extract text from an Excel file.
    Args:
        file_bytes (bytes): The content of the Excel file in bytes.
        max_rows (int): Maximum number of rows to extract per sheet.
    Returns:
        str: Extracted text content.
    """
    bio = io.BytesIO(file_bytes)
    wb = load_workbook(bio, data_only=True)

    sheet_texts = []

    for sheet in wb.worksheets:
        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0

        if max_row == 0 or max_col == 0:
            continue

        # get last rows to read
        last_row = min(max_row, max_rows)

        # Header of Excel: A,B,C,...
        col_letters = [get_column_letter(c) for c in range(1, max_col + 1)]
        rows_str = []

        # first row: header
        header_line = ";".join(["ROW"] + col_letters)
        rows_str.append(header_line)

        # rows with data and empty cells as ""
        for r in range(1, last_row + 1):
            row_values = []
            for c in range(1, max_col + 1):
                cell = sheet.cell(row=r, column=c)
                val = cell.value
                if val is None:
                    row_values.append("")  # cella vuota
                else:
                    row_values.append(str(val))
            # prima colonna: numero di riga
            row_line = ";".join([str(r)] + row_values)
            rows_str.append(row_line)

        sheet_block = f"--- Sheet: {sheet.title} ---\n" + "\n".join(rows_str)
        sheet_texts.append(sheet_block)

    if not sheet_texts:
        return "File Excel ricevuto ma non è stato possibile estrarre testo o non contiene dati."

    return "\n\n".join(sheet_texts)


def extract_text_from_csv(file_bytes: bytes, max_rows: int = 50) -> str:
    """
    Extract text from a CSV file.
    Args:
        file_bytes (bytes): The content of the CSV file in bytes.
        max_rows (int): Maximum number of rows to extract.
    Returns:
        str: Extracted text content."""
    decoded = file_bytes.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(decoded))

    rows = []
    for i, row in enumerate(reader):
        if i >= max_rows:
            break
        rows.append(", ".join(row))

    if not rows:
        return "File CSV ricevuto ma non è stato possibile estrarre testo."
    return "\n".join(rows)


def split_plain(text: str, max_len=4000):
    """
    Divide text in chunk with max_len characters, trying to split at newlines or spaces.
    Args:
        text (str): The input text to split.
        max_len (int): Maximum length of each chunk.
    Returns:
        List[str]: List of text chunks.
    """
    parts = []
    while len(text) > max_len:
        split_at = text.rfind("\n", 0, max_len)
        if split_at == -1:
            split_at = text.rfind(" ", 0, max_len)
        if split_at == -1:
            split_at = max_len
        parts.append(text[:split_at])
        text = text[split_at:]
    if text:
        parts.append(text)
    return parts


def split_code_block(code: str, max_len=4000):
    """
    Divide code block in chunk with max_len characters, preserving code block formatting.
    Args:
        code (str): The input code to split.
        max_len (int): Maximum length of each chunk.
    Returns:
        List[str]: List of code block chunks.
    """
    parts = []
    while len(code) > max_len:
        parts.append("```\n" + code[:max_len] + "\n```")
        code = code[max_len:]
    if code:
        parts.append("```\n" + code + "\n```")
    return parts


def split_markdown(text: str, max_len=4000):
    """
    Divide markdownv2 text in chunk with max_len characters, preserving code blocks.
    Args:
        text (str): The input markdownv2 text to split.
        max_len (int): Maximum length of each chunk.
    Returns:
        List[str]: List of markdownv2 text chunks.
    """
    parts = []
    segments = re.split(r"(```.*?```)", text, flags=re.DOTALL)

    for seg in segments:
        if seg.startswith("```") and seg.endswith("```"):
            code = seg[3:-3].strip("\n")  # estrai contenuto senza ```
            parts.extend(split_code_block(code, max_len))
        else:
            parts.extend(split_plain(seg, max_len))

    return parts
